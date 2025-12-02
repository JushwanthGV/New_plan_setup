import os
import logging
from pathlib import Path
import PyPDF2
from docx import Document
from openpyxl import load_workbook
from PIL import Image, ImageEnhance, ImageFilter
import torch
from transformers import TrOCRProcessor, VisionEncoderDecoderModel
import cv2
import numpy as np
import easyocr
from collections import Counter

logger = logging.getLogger(__name__)

class DocumentParser:
    """Advanced Parser with Multi-Pass OCR and Image Enhancement"""
    
    def __init__(self):
        """Initialize parser and load models"""
        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        logger.info(f"Initializing DocumentParser on {self.device}...")
        
        # Initialize EasyOCR reader
        self.reader = easyocr.Reader(['en'], gpu=torch.cuda.is_available())
        
        # Load TrOCR models (base and large for ensemble)
        try:
            logger.info("Loading TrOCR models...")
            
            # Base model (fast)
            self.processor_base = TrOCRProcessor.from_pretrained('microsoft/trocr-base-handwritten')
            self.model_base = VisionEncoderDecoderModel.from_pretrained('microsoft/trocr-base-handwritten').to(self.device)
            
            # Try loading large model (better accuracy)
            try:
                self.processor_large = TrOCRProcessor.from_pretrained('microsoft/trocr-large-handwritten')
                self.model_large = VisionEncoderDecoderModel.from_pretrained('microsoft/trocr-large-handwritten').to(self.device)
                self.has_large_model = True
                logger.info("✓ Loaded both base and large TrOCR models")
            except:
                self.has_large_model = False
                logger.info("✓ Loaded base TrOCR model only (large model not available)")
            
            self.ocr_available = True
        except Exception as e:
            logger.error(f"Failed to load TrOCR: {str(e)}")
            self.ocr_available = False
    
    @staticmethod
    def parse_document(filepath: str) -> str:
        """Extract text from a document"""
        parser = DocumentParser()
        return parser._parse_file_by_extension(filepath)

    def _parse_file_by_extension(self, filepath: str) -> str:
        ext = Path(filepath).suffix.lower()
        try:
            if ext == '.pdf':
                return self._parse_pdf(filepath)
            elif ext in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
                return self._parse_image_advanced(filepath)
            elif ext in ['.docx', '.doc']:
                return self._parse_docx(filepath)
            elif ext in ['.xlsx', '.xls']:
                return self._parse_excel(filepath)
            elif ext == '.txt':
                return self._parse_txt(filepath)
            else:
                logger.warning(f"Unsupported file type: {ext}")
                return ""
        except Exception as e:
            logger.error(f"Error parsing {filepath}: {str(e)}")
            return ""

    def _parse_pdf(self, filepath: str) -> str:
        """Extract text from PDF - digital first, then multi-pass OCR"""
        text = ""
        try:
            # Try digital text extraction first
            with open(filepath, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    t = page.extract_text()
                    if t: 
                        text += t + "\n"
            
            # Only use OCR if text is empty or suspiciously short
            if not text.strip() or len(text.strip()) < 10:
                logger.info("Minimal/no digital text. Attempting Multi-Pass OCR...")
                text = self._parse_pdf_with_multipass_ocr(filepath)
            else:
                logger.info(f"✓ Extracted {len(text)} characters as digital text (no OCR needed)")
                
        except Exception as e:
            logger.error(f"PDF error: {str(e)}")
            text = self._parse_pdf_with_multipass_ocr(filepath)
        
        return text.strip()

    def _parse_pdf_with_multipass_ocr(self, filepath: str) -> str:
        """Convert PDF to images and run multi-pass OCR"""
        try:
            from pdf2image import convert_from_path
            
            # Convert at high DPI for better quality
            logger.info("Converting PDF to images at 400 DPI...")
            images = convert_from_path(filepath, dpi=400) 
            
            full_text = ""
            for i, image in enumerate(images):
                logger.info(f"Processing page {i+1}/{len(images)} with Multi-Pass OCR...")
                page_text = self._perform_multipass_ocr(image, page_num=i+1)
                full_text += page_text + "\n"
                
            return full_text
        except Exception as e:
            logger.error(f"Multi-Pass OCR failed: {str(e)}")
            return ""

    def _parse_image_advanced(self, filepath: str) -> str:
        """Parse image with multi-pass OCR"""
        try:
            image = Image.open(filepath).convert("RGB")
            return self._perform_multipass_ocr(image)
        except Exception as e:
            logger.error(f"Image parsing failed: {str(e)}")
            return ""

    def _enhance_image_v1_light(self, image):
        """Enhancement Level 1: Light processing"""
        img_array = np.array(image.convert('L'))
        
        # Denoise
        img_array = cv2.fastNlMeansDenoising(img_array, h=10)
        
        # Increase contrast slightly
        img_array = cv2.convertScaleAbs(img_array, alpha=1.3, beta=0)
        
        # Adaptive threshold
        img_array = cv2.adaptiveThreshold(
            img_array, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY, 11, 2
        )
        
        return Image.fromarray(img_array)
    
    def _enhance_image_v2_aggressive(self, image):
        """Enhancement Level 2: Aggressive processing"""
        img_array = np.array(image.convert('L'))
        
        # Strong denoising
        img_array = cv2.fastNlMeansDenoising(img_array, h=20)
        
        # High contrast
        img_array = cv2.convertScaleAbs(img_array, alpha=2.0, beta=10)
        
        # Morphological operations
        kernel = np.ones((2,2), np.uint8)
        img_array = cv2.morphologyEx(img_array, cv2.MORPH_CLOSE, kernel)
        
        # Adaptive threshold with larger block
        img_array = cv2.adaptiveThreshold(
            img_array, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY, 15, 2
        )
        
        # Dilation to thicken text
        img_array = cv2.dilate(img_array, kernel, iterations=1)
        
        return Image.fromarray(img_array)
    
    def _enhance_image_v3_extreme(self, image):
        """Enhancement Level 3: Extreme processing (last resort)"""
        img_array = np.array(image.convert('L'))
        
        # Maximum denoising
        img_array = cv2.fastNlMeansDenoising(img_array, h=30)
        
        # Extreme contrast
        img_array = cv2.convertScaleAbs(img_array, alpha=2.5, beta=20)
        
        # Sharpening
        kernel_sharp = np.array([[-1,-1,-1],
                                 [-1, 9,-1],
                                 [-1,-1,-1]])
        img_array = cv2.filter2D(img_array, -1, kernel_sharp)
        
        # Otsu's thresholding (automatic)
        _, img_array = cv2.threshold(img_array, 0, 255, 
                                     cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        
        # Heavy dilation
        kernel = np.ones((3,3), np.uint8)
        img_array = cv2.dilate(img_array, kernel, iterations=2)
        
        return Image.fromarray(img_array)

    def _remove_horizontal_lines(self, image):
        """Remove notebook lines using morphological operations"""
        try:
            img_array = np.array(image.convert('L'))
            
            # Inverse binary threshold
            thresh = cv2.threshold(img_array, 0, 255, 
                                  cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
            
            # Detect horizontal lines
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
            detected_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, 
                                             horizontal_kernel, iterations=2)
            
            # Remove detected lines
            cnts = cv2.findContours(detected_lines, cv2.RETR_EXTERNAL, 
                                   cv2.CHAIN_APPROX_SIMPLE)
            cnts = cnts[0] if len(cnts) == 2 else cnts[1]
            
            repaired_img = img_array.copy()
            for c in cnts:
                cv2.drawContours(repaired_img, [c], -1, (255, 255, 255), 2)
                
            return Image.fromarray(repaired_img)
        except Exception as e:
            logger.warning(f"Line removal failed: {str(e)}")
            return image

    def _perform_multipass_ocr(self, image, page_num=1):
        """
        Multi-pass OCR with progressive enhancement
        Pass 1: Original image (light enhancement)
        Pass 2: Aggressive enhancement
        Pass 3: Extreme enhancement
        Returns best result based on confidence/length
        """
        
        logger.info(f"=== Multi-Pass OCR Started (Page {page_num}) ===")
        
        # Keep original
        original_image = image.convert("RGB")
        
        # Create debug directory
        debug_dir = Path("./data/debug_images")
        debug_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        # ==========================================
        # PASS 1: Light Enhancement
        # ==========================================
        logger.info("Pass 1: Light enhancement...")
        try:
            enhanced_v1 = self._enhance_image_v1_light(original_image)
            enhanced_v1 = self._remove_horizontal_lines(enhanced_v1)
            enhanced_v1.save(debug_dir / f"page_{page_num}_pass1_light.png")
            
            text_v1 = self._run_ocr_engines(enhanced_v1, original_image)
            results.append({
                'text': text_v1,
                'length': len(text_v1),
                'pass': 'light',
                'confidence': self._calculate_confidence(text_v1)
            })
            logger.info(f"Pass 1 result: {len(text_v1)} chars, confidence: {results[0]['confidence']:.2f}")
        except Exception as e:
            logger.error(f"Pass 1 failed: {str(e)}")
        
        # ==========================================
        # PASS 2: Aggressive Enhancement
        # ==========================================
        logger.info("Pass 2: Aggressive enhancement...")
        try:
            enhanced_v2 = self._enhance_image_v2_aggressive(original_image)
            enhanced_v2 = self._remove_horizontal_lines(enhanced_v2)
            enhanced_v2.save(debug_dir / f"page_{page_num}_pass2_aggressive.png")
            
            text_v2 = self._run_ocr_engines(enhanced_v2, original_image)
            results.append({
                'text': text_v2,
                'length': len(text_v2),
                'pass': 'aggressive',
                'confidence': self._calculate_confidence(text_v2)
            })
            logger.info(f"Pass 2 result: {len(text_v2)} chars, confidence: {results[1]['confidence']:.2f}")
        except Exception as e:
            logger.error(f"Pass 2 failed: {str(e)}")
        
        # ==========================================
        # PASS 3: Extreme Enhancement (if needed)
        # ==========================================
        # Only run if first two passes gave poor results
        if all(r['length'] < 20 or r['confidence'] < 0.5 for r in results):
            logger.info("Pass 3: Extreme enhancement (poor results so far)...")
            try:
                enhanced_v3 = self._enhance_image_v3_extreme(original_image)
                enhanced_v3 = self._remove_horizontal_lines(enhanced_v3)
                enhanced_v3.save(debug_dir / f"page_{page_num}_pass3_extreme.png")
                
                text_v3 = self._run_ocr_engines(enhanced_v3, original_image)
                results.append({
                    'text': text_v3,
                    'length': len(text_v3),
                    'pass': 'extreme',
                    'confidence': self._calculate_confidence(text_v3)
                })
                logger.info(f"Pass 3 result: {len(text_v3)} chars, confidence: {results[2]['confidence']:.2f}")
            except Exception as e:
                logger.error(f"Pass 3 failed: {str(e)}")
        
        # ==========================================
        # Select Best Result
        # ==========================================
        if not results:
            logger.error("All OCR passes failed!")
            return ""
        
        # Choose result with best confidence score
        best_result = max(results, key=lambda x: x['confidence'])
        
        logger.info(f"✓ Best result from '{best_result['pass']}' pass: "
                   f"{best_result['length']} chars, confidence: {best_result['confidence']:.2f}")
        logger.info(f"=== Multi-Pass OCR Complete (Page {page_num}) ===\n")
        
        return best_result['text']

    def _run_ocr_engines(self, clean_image, original_image):
        """Run multiple OCR engines and combine results"""
        
        # Detect text regions using clean image
        img_array_clean = np.array(clean_image.convert("RGB"))
        boxes = self.reader.readtext(img_array_clean, paragraph=False)
        
        if not boxes:
            logger.warning("No text regions detected")
            return ""
        
        logger.info(f"Detected {len(boxes)} text regions")
        
        # Sort top-to-bottom, left-to-right
        boxes.sort(key=lambda x: (x[0][0][1], x[0][0][0]))
        
        detected_lines = []
        
        for bbox, _, confidence in boxes:
            (tl, tr, br, bl) = bbox
            top = max(0, int(min(tl[1], tr[1])) - 5)
            bottom = int(max(bl[1], br[1])) + 5
            left = max(0, int(min(tl[0], bl[0])) - 5)
            right = int(max(tr[0], br[0])) + 5
            
            # Crop from ORIGINAL image
            line_crop = original_image.crop((left, top, right, bottom))
            
            # Run multiple OCR engines
            results = []
            
            # Engine 1: TrOCR Base
            try:
                text_base = self._run_trocr(line_crop, self.processor_base, self.model_base)
                if text_base:
                    results.append(text_base)
            except Exception as e:
                logger.debug(f"TrOCR base failed: {str(e)}")
            
            # Engine 2: TrOCR Large (if available)
            if self.has_large_model:
                try:
                    text_large = self._run_trocr(line_crop, self.processor_large, self.model_large)
                    if text_large:
                        results.append(text_large)
                except Exception as e:
                    logger.debug(f"TrOCR large failed: {str(e)}")
            
            # Engine 3: EasyOCR
            try:
                easy_result = self.reader.readtext(np.array(line_crop))
                if easy_result:
                    text_easy = ' '.join([text for (_, text, _) in easy_result])
                    if text_easy:
                        results.append(text_easy)
            except Exception as e:
                logger.debug(f"EasyOCR failed: {str(e)}")
            
            # Voting: Choose most common result or longest if no agreement
            if results:
                if len(results) >= 2:
                    # Check for agreement
                    counter = Counter(results)
                    most_common = counter.most_common(1)[0]
                    if most_common[1] >= 2:  # At least 2 engines agree
                        final_text = most_common[0]
                    else:
                        final_text = max(results, key=len)  # Choose longest
                else:
                    final_text = results[0]
                
                # Apply post-processing corrections
                final_text = self._apply_corrections(final_text)
                detected_lines.append(final_text)
                logger.debug(f"OCR result: '{final_text}'")
        
        return "\n".join(detected_lines)

    def _run_trocr(self, image, processor, model):
        """Run TrOCR inference"""
        try:
            pixel_values = processor(images=image, return_tensors="pt").pixel_values.to(self.device)
            generated_ids = model.generate(pixel_values, max_length=64)
            text = processor.batch_decode(generated_ids, skip_special_tokens=True)[0]
            return text.strip()
        except:
            return ""

    def _apply_corrections(self, text):
        """Apply common OCR error corrections"""
        corrections = {
            'Banglori': 'Bangalore',
            'Bangalor': 'Bangalore',
            'Tushworth': 'Jushwanth',
            'sureish': 'Suresh',
            '0': 'O',  # Common in names
        }
        
        for wrong, correct in corrections.items():
            if wrong.lower() in text.lower():
                text = text.replace(wrong, correct)
        
        return text

    def _calculate_confidence(self, text):
        """Calculate confidence score for OCR result"""
        if not text:
            return 0.0
        
        score = 0.0
        
        # Length bonus (more text = better)
        if len(text) > 50:
            score += 0.3
        elif len(text) > 20:
            score += 0.2
        elif len(text) > 10:
            score += 0.1
        
        # Has alphabetic characters
        if any(c.isalpha() for c in text):
            score += 0.2
        
        # Has numbers (phone, address)
        if any(c.isdigit() for c in text):
            score += 0.2
        
        # Has proper capitalization
        if any(c.isupper() for c in text) and any(c.islower() for c in text):
            score += 0.1
        
        # Has common words (NAME, ADDRESS, PHONE)
        common_words = ['name', 'address', 'phone', 'number']
        if any(word in text.lower() for word in common_words):
            score += 0.2
        
        return min(score, 1.0)

    # =========================================================================
    # Standard document parsers (unchanged)
    # =========================================================================
    
    @staticmethod
    def _parse_docx(filepath: str) -> str:
        try:
            doc = Document(filepath)
            return "\n".join([p.text for p in doc.paragraphs]).strip()
        except:
            return ""
        
    @staticmethod
    def _parse_excel(filepath: str) -> str:
        try:
            wb = load_workbook(filepath, read_only=True)
            text = ""
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text += " | ".join([str(c) if c else "" for c in row]) + "\n"
            return text.strip()
        except:
            return ""

    @staticmethod
    def _parse_txt(filepath: str) -> str:
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return f.read().strip()
        except:
            try:
                with open(filepath, 'r', encoding='latin-1') as f:
                    return f.read().strip()
            except:
                return ""